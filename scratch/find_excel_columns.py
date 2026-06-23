import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    # We want to print all headers on row 12
    headers = {}
    for col in range(1, 140):
        val = sheet.cell(row=12, column=col).value
        parent_val = sheet.cell(row=11, column=col).value
        grand_val = sheet.cell(row=10, column=col).value
        
        if val is not None or parent_val is not None:
            headers[col] = (grand_val, parent_val, val)
            
    print("Column headers in sheet:")
    for col, val in headers.items():
        print(f"Col {col}: Grand={val[0]} | Parent={val[1]} | Val={val[2]}")
except Exception as e:
    print("Error:", e)
