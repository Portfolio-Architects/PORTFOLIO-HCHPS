import os
import sys
import re
import zipfile
import json
import argparse
from typing import Dict, List, Any

# HWPX, PDF, XLSX, TXT 파싱 라이브러리 임포트 가드
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Windows UTF-8 입출력 강제
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# 검색 타겟 확장자
ALLOWED_EXTENSIONS = {'.pdf', '.hwpx', '.xlsx', '.xls', '.txt', '.md', '.csv', '.json'}
CACHE_FILENAME = ".search_cache.json"

def load_cache(archive_path: str) -> dict:
    """검색 경로 내의 텍스트 본문 캐시 데이터를 로드"""
    cache_path = os.path.join(archive_path, CACHE_FILENAME)
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_cache(archive_path: str, cache_data: dict):
    """텍스트 본문 캐시 데이터를 파일에 저장"""
    cache_path = os.path.join(archive_path, CACHE_FILENAME)
    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        sys.stderr.write(f"[캐시 저장 경고] {e}\n")

def parse_pdf(filepath: str) -> str:
    """PDF 본문 텍스트 추출"""
    if not fitz:
        return "[PyMuPDF(fitz) 라이브러리가 설치되지 않아 PDF 파싱 불가]"
    text_list = []
    try:
        doc = fitz.open(filepath)
        for page in doc:
            text_list.append(page.get_text())
        return "\n".join(text_list)
    except Exception as e:
        return f"[PDF 파싱 오류: {e}]"

def parse_hwpx(filepath: str) -> str:
    """HWPX 본문 텍스트 추출 (XML 구조 분석)"""
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            sections = [f for f in z.namelist() if f.startswith('Contents/section') and f.endswith('.xml')]
            text_list = []
            for section in sections:
                xml_content = z.read(section).decode('utf-8', errors='ignore')
                text = re.sub(r'<[^>]+>', ' ', xml_content)
                text = re.sub(r'\s+', ' ', text)
                text_list.append(text)
            return "\n".join(text_list)
    except Exception as e:
        return f"[HWPX 파싱 오류: {e}]"

def parse_xlsx(filepath: str) -> str:
    """XLSX 엑셀 시트 본문 텍스트 추출"""
    if not openpyxl:
        return "[openpyxl 라이브러리가 설치되지 않아 Excel 파싱 불가]"
    all_text = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                row_str = " ".join([str(val) for val in row if val is not None])
                if row_str.strip():
                    all_text.append(row_str)
        return "\n".join(all_text)
    except Exception as e:
        return f"[Excel 파싱 오류: {e}]"

def parse_text(filepath: str) -> str:
    """일반 텍스트(TXT, MD, CSV 등) 추출"""
    encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc, errors='ignore') as f:
                return f.read()
        except Exception:
            continue
    return "[텍스트 인코딩 오류]"

def get_file_content(filepath: str) -> str:
    """확장자에 따른 본문 통합 추출기"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.pdf':
        return parse_pdf(filepath)
    elif ext == '.hwpx':
        return parse_hwpx(filepath)
    elif ext == '.xlsx' or ext == '.xls':
        return parse_xlsx(filepath)
    elif ext in {'.txt', '.md', '.csv', '.json'}:
        return parse_text(filepath)
    return ""

def search_in_archives(search_path: str, query: str, is_json: bool = False) -> List[Dict[str, Any]]:
    """지정된 디렉토리 내 모든 문서를 캐시 기반으로 초고속 탐색하여 키워드 매칭 수행"""
    matches = []
    
    if not os.path.exists(search_path):
        if not is_json:
            print(f"[오류] 대상 경로가 존재하지 않습니다: {search_path}")
        return []

    # 1. 캐시 로드
    cache_data = load_cache(search_path)
    cache_updated = False

    if not is_json:
        print(f"🔍 '{search_path}' 경로 내에서 본문 검색 시작: '{query}'")
        print("점진적 캐시 분석 및 메모리 고속 매칭을 수행하는 중...")

    for root, _, files in os.walk(search_path):
        for file in files:
            # 캐시 파일 자체는 스캔에서 배제
            if file == CACHE_FILENAME:
                continue

            filepath = os.path.join(root, file)
            ext = os.path.splitext(file)[1].lower()
            if ext not in ALLOWED_EXTENSIONS:
                continue

            try:
                stat = os.stat(filepath)
                mtime = stat.st_mtime
                size = stat.st_size
            except Exception:
                continue

            content = ""
            # 캐시 히트 검사 (경로, 최종 수정시각, 파일 크기가 모두 완벽히 동일할 경우)
            if filepath in cache_data:
                file_cache = cache_data[filepath]
                if file_cache.get("mtime") == mtime and file_cache.get("size") == size:
                    content = file_cache.get("content", "")

            # 캐시 미스 또는 정보 변경 시 새로 파싱하고 캐시 갱신
            if not content:
                content = get_file_content(filepath)
                if content:
                    cache_data[filepath] = {
                        "mtime": mtime,
                        "size": size,
                        "content": content
                    }
                    cache_updated = True

            if not content:
                continue

            pattern = re.compile(re.escape(query), re.IGNORECASE)
            finds = list(pattern.finditer(content))

            if finds:
                file_matches = []
                for match in finds[:5]:  # 최대 5개의 스니펫만 추출
                    start = max(0, match.start() - 60)
                    end = min(len(content), match.end() + 60)
                    snippet = content[start:end].replace('\n', ' ').strip()
                    file_matches.append({
                        "pos": match.start(),
                        "snippet": f"... {snippet} ..."
                    })
                
                matches.append({
                    "fileName": file,
                    "relPath": os.path.relpath(filepath, search_path),
                    "fullPath": filepath,
                    "count": len(finds),
                    "snippets": file_matches
                })
                
    # 2. 캐시 변경 내역이 있을 시 캐시 파일 갱신
    if cache_updated:
        save_cache(search_path, cache_data)
        
    return matches

def main():
    parser = argparse.ArgumentParser(description="로컬 문서 본문 고속 텍스트 검색 툴")
    parser.add_argument("-q", "--query", type=str, help="검색할 본문 키워드")
    parser.add_argument("-p", "--path", type=str, default=r"F:\부엉이_정리됨", help="탐색할 아카이브 루트 경로")
    parser.add_argument("-j", "--json", action="store_true", help="결과를 JSON 형식으로만 출력")
    args = parser.parse_args()

    query = args.query
    if not query and not args.json:
        query = input("🔎 검색할 문서 내부 키워드를 입력하세요: ").strip()

    if not query:
        if args.json:
            print("[]")
        else:
            print("[오류] 검색어가 비어 있습니다.")
        sys.exit(1)

    results = search_in_archives(args.path, query, args.json)

    if args.json:
        # JSON 포맷으로만 출력
        print(json.dumps(results, ensure_ascii=False, indent=2))
        return

    if not results:
        print(f"❌ '{query}' 키워드가 포함된 문서를 찾지 못했습니다.")
        return

    print(f"🎉 총 {len(results)}개의 파일에서 '{query}' 키워드를 발견했습니다:\n")
    for idx, res in enumerate(results, 1):
        print(f"[{idx}] {res['fileName']} (매칭 횟수: {res['count']}회)")
        print(f"   📂 경로: {res['fullPath']}")
        print("   📝 매칭 문맥(일부):")
        for snip in res['snippets']:
            print(f"      {snip['snippet']}")
        print("-" * 80)

if __name__ == "__main__":
    main()
