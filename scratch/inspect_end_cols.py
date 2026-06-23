import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    print("Inspecting columns from 138 to 150:")
    for col in range(138, 151):
        if col > sheet.max_column:
            break
        letter = openpyxl.utils.get_column_letter(col)
        h8 = sheet.cell(row=8, column=col).value
        h9 = sheet.cell(row=9, column=col).value
        h10 = sheet.cell(row=10, column=col).value
        h11 = sheet.cell(row=11, column=col).value
        print(f"Col {col:03d} ({letter}): Row8='{h8}', Row9='{h9}', Row10='{h10}', Row11='{h11}'")
        
        # Let's print the value in row 13 (first participant) for this column
        p_val = sheet.cell(row=13, column=col).value
        print(f"  Row 13 value: {p_val}")
        
except Exception as e:
    print(f"Error: {e}")
