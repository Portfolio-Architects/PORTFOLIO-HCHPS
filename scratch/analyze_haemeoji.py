import sys
import openpyxl
import os
import math

sys.stdout.reconfigure(encoding='utf-8')

desktop_dir = r"D:\Desktop"
meas_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx")
att_file = os.path.join(desktop_dir, "7. 체력향상 프로그램 출석부_2026_강남구(해맞이).xlsx")

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return None

def calculate_paired_t_test(pre_list, post_list):
    n = len(pre_list)
    if n < 2:
        return None, None, None, None, None
        
    diffs = [post - pre for pre, post in zip(pre_list, post_list)]
    mean_diff = sum(diffs) / n
    
    # Standard deviation of differences
    sq_diff_sum = sum((d - mean_diff) ** 2 for d in diffs)
    std_dev = math.sqrt(sq_diff_sum / (n - 1))
    
    # Standard error
    std_err = std_dev / math.sqrt(n)
    
    # T-statistic
    if std_err == 0:
        t_stat = 0.0
    else:
        t_stat = mean_diff / std_err
        
    df = n - 1
    t_abs = abs(t_stat)
    
    try:
        theta = math.atan(t_abs / math.sqrt(df))
        if df == 1:
            p_val = 1.0 - (2.0 / math.pi) * theta
        elif df == 2:
            p_val = 1.0 - t_abs / math.sqrt(t_abs**2 + 2)
        elif df == 3:
            p_val = 1.0 - (2.0 / math.pi) * (theta + t_abs * math.sqrt(df) / (t_abs**2 + df))
        elif df == 4:
            x = t_abs / math.sqrt(t_abs**2 + 4)
            p_val = 2.0 * (1.0 - (0.5 + 0.5 * (x + 0.5 * x * (1.0 - x**2))))
        else:
            # General approximation
            z = t_abs * (1.0 - 1.0 / (4.0 * df)) / math.sqrt(1.0 + t_abs**2 / (2.0 * df))
            p_val = math.erfc(z / math.sqrt(2.0))
    except Exception:
        p_val = None
        
    return mean_diff, std_dev, std_err, t_stat, p_val

try:
    print("====================================================================")
    print("■ 삼성 해맞이 공원 프로그램 데이터 분석")
    print("====================================================================")
    
    # Load Measurement File
    wb_meas = openpyxl.load_workbook(meas_file, data_only=True)
    sheet_meas = wb_meas['사전사후측정결과(출석부 포함)']
    
    district = sheet_meas['B2'].value
    location = sheet_meas['B3'].value
    period = sheet_meas['B4'].value
    day_time = sheet_meas['B5'].value
    registered_count = sheet_meas['B6'].value
    instructor = sheet_meas['B7'].value
    
    print("\n[프로그램 개요]")
    print(f"- 자치구: {district}")
    print(f"- 운영장소: {location}")
    print(f"- 운영기간: {period}")
    print(f"- 운영요일/시간: {day_time}")
    print(f"- 등록인원: {registered_count}")
    print(f"- 강사명/소속: {instructor}")
    
    # Parse Participant Details
    participants = []
    completed_participants = []
    
    for row_idx in range(13, 212):
        row = [sheet_meas.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3] # Column D
        if not name:
            continue
            
        serial_no = row[1] # Column B
        age = row[5] # Column F
        
        # Pre-measurements
        pre_w = safe_float(row[7])
        pre_m = safe_float(row[8])
        pre_f = safe_float(row[9])
        pre_bmi = safe_float(row[10])
        pre_waist = safe_float(row[11])
        
        # Post-measurements
        post_w = safe_float(row[13])
        post_m = safe_float(row[14])
        post_f = safe_float(row[15])
        post_bmi = safe_float(row[16])
        post_waist = safe_float(row[17])
        
        # GPAQ
        gpaq_pre = safe_float(row[40]) if row[40] is not None else 0.0
        gpaq_post = safe_float(row[66]) if row[66] is not None else 0.0
        
        p_data = {
            'no': serial_no,
            'name': name,
            'age': age,
            'pre': {'weight': pre_w, 'muscle': pre_m, 'fat': pre_f, 'bmi': pre_bmi, 'waist': pre_waist, 'gpaq': gpaq_pre},
            'post': {'weight': post_w, 'muscle': post_m, 'fat': post_f, 'bmi': post_bmi, 'waist': post_waist, 'gpaq': gpaq_post}
        }
        participants.append(p_data)
        
        if post_w is not None:
            completed_participants.append(p_data)
            
    print(f"\n[등록 회원 측정 결과 (총 {len(participants)}명 중 완수자 {len(completed_participants)}명)]")
    print("=" * 110)
    print(f"{'연번':<4}{'이름':<6}{'나이':<4}{'사전 체중/근육/지방':<22}{'사후 체중/근육/지방':<22}{'체중 변화':<10}{'골격근 변화':<10}{'지방률 변화':<10}")
    print("-" * 110)
    for p in participants:
        pre_str = f"{p['pre']['weight'] or 0:.1f} / {p['pre']['muscle'] or 0:.1f} / {p['pre']['fat'] or 0:.1f}"
        post_str = f"{p['post']['weight'] or 0:.1f} / {p['post']['muscle'] or 0:.1f} / {p['post']['fat'] or 0:.1f}" if p['post']['weight'] else "측정없음"
        
        # Changes
        if p['post']['weight'] and p['pre']['weight']:
            w_diff = p['post']['weight'] - p['pre']['weight']
            w_diff_str = f"{w_diff:+.1f} kg"
        else:
            w_diff_str = "-"
            
        if p['post']['muscle'] and p['pre']['muscle']:
            m_diff = p['post']['muscle'] - p['pre']['muscle']
            m_diff_str = f"{m_diff:+.1f} kg"
        else:
            m_diff_str = "-"
            
        if p['post']['fat'] and p['pre']['fat']:
            f_diff = p['post']['fat'] - p['pre']['fat']
            f_diff_str = f"{f_diff:+.1f}%"
        else:
            f_diff_str = "-"
            
        print(f"{p['no']:<4}{p['name']:<6}{p['age']:<4}{pre_str:<25}{post_str:<25}{w_diff_str:<12}{m_diff_str:<12}{f_diff_str:<12}")
    print("=" * 110)

    # Parse Attendance Log File
    wb_att = openpyxl.load_workbook(att_file, data_only=True)
    sheet_att = wb_att[wb_att.sheetnames[0]]
    
    attendance_data = {}
    print(f"\n[실제 출석 집계 ({att_file.split(os.sep)[-1]})]")
    print("-" * 60)
    for r in range(4, 40):
        name = sheet_att.cell(row=r, column=3).value
        if not name:
            continue
        serial = sheet_att.cell(row=r, column=2).value
        
        attended_count = 0
        for c in range(4, 20):  # columns D to S (16 sessions)
            val = sheet_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended_count += 1
        
        rate = (attended_count / 16.0) * 100.0
        attendance_data[name] = {'attended': attended_count, 'rate': rate}
        print(f"연번 {serial:<3} {name:<6} : {attended_count:>2} / 16회 출석 ({rate:.1f}%)")
        
    # Parse Satisfaction Survey
    sheet_survey = wb_meas['만족도조사결과']
    survey_rows = list(sheet_survey.iter_rows(values_only=True))
    
    questions = [cell for cell in survey_rows[1] if cell is not None]
    
    respondents = []
    q_sums = [0.0] * 11
    q_counts = [0] * 11
    comments = []
    
    for row in survey_rows[2:]:
        if not row or row[0] is None:
            continue
        serial = row[0]
        scores = []
        for i in range(1, 12):
            val = row[i]
            if val is not None:
                score = safe_float(val)
                if score is not None:
                    q_sums[i-1] += score
                    q_counts[i-1] += 1
                    scores.append(score)
                else:
                    scores.append(None)
            else:
                scores.append(None)
        
        comment = row[12] if len(row) > 12 else None
        respondents.append({
            'no': serial,
            'scores': scores,
            'comment': comment
        })
        if comment:
            comments.append(f"[{serial}번 회원] {str(comment).strip()}")
            
    print(f"\n[사후 만족도 조사 결과 (총 {len(respondents)}명 응답)]")
    print("-" * 80)
    for idx in range(11):
        q_label = questions[idx+1]
        avg_score = q_sums[idx] / q_counts[idx] if q_counts[idx] > 0 else 0.0
        print(f"Q{idx+1:02d}. {q_label[:55]}... => 평균: {avg_score:.2f}점")
    
    print("\n[회원 기타 의견 및 건의 사항 (Q12)]")
    for c in comments:
        print(c)
        
    # Paired t-test
    print(f"\n[대응표본 t-검정 (Paired t-test) 분석 (완수자 N = {len(completed_participants)}, df = {len(completed_participants)-1})]")
    print("=" * 110)
    print(f"{'평가 지표':<15}{'사전 평균':<12}{'사후 평균':<12}{'평균 변화량':<12}{'표준편차(diff)':<15}{'t-통계량':<12}{'p-value':<12}{'유의성 (p < 0.05)':<15}")
    print("-" * 110)
    
    metrics_list = ['weight', 'muscle', 'fat', 'waist', 'gpaq']
    for m in metrics_list:
        pre_vals = []
        post_vals = []
        for p in completed_participants:
            pre_val = p['pre'][m]
            post_val = p['post'][m]
            if pre_val is not None and post_val is not None:
                pre_vals.append(pre_val)
                post_vals.append(post_val)
                
        if len(pre_vals) >= 2:
            avg_pre = sum(pre_vals) / len(pre_vals)
            avg_post = sum(post_vals) / len(post_vals)
            mean_diff, std_dev, std_err, t_stat, p_val = calculate_paired_t_test(pre_vals, post_vals)
            
            sig_str = "유의함 (Significant)" if p_val is not None and p_val < 0.05 else "유의하지 않음"
            p_val_str = f"{p_val:.4f}" if p_val is not None else "N/A"
            t_stat_str = f"{t_stat:+.3f}" if t_stat is not None else "N/A"
            mean_diff_str = f"{mean_diff:+.2f}"
            print(f"{m.capitalize():<15}{avg_pre:<15.2f}{avg_post:<15.2f}{mean_diff_str:<15}{std_dev:<15.4f}{t_stat_str:<12}{p_val_str:<12}{sig_str:<15}")
        else:
            print(f"{m.capitalize():<15} - 데이터 부족 (완수자 N={len(pre_vals)})")
            
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
