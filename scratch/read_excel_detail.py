import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\사업운영 관련 연락처.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("=== 건강증진지원실 전체 행 ===")
sheet = wb['건강증진지원실']
for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if any(cell is not None for cell in row):
        cleaned_row = [str(cell) if cell is not None else '' for cell in row]
        print(f"Row {r_idx+1:02d}: {cleaned_row}")

print("\n=== 어린이 성장발달 시스템 전체 행 ===")
sheet = wb['어린이 성장발달 시스템']
for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if any(cell is not None for cell in row):
        cleaned_row = [str(cell) if cell is not None else '' for cell in row]
        print(f"Row {r_idx+1:02d}: {cleaned_row}")
