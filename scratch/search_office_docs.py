import os
import sys
import zipfile
import re
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
keywords = ['체질량', '수축기', '이완기', '머리척추각', 'CVA', '둘레', '혈압', 'BMI']

def search_text(text, filename):
    for kw in keywords:
        matches = list(re.finditer(re.escape(kw), text, re.IGNORECASE))
        if matches:
            print(f"[FOUND] '{kw}' in file: {filename} ({len(matches)} times)")
            for match in matches[:3]:
                start = max(0, match.start() - 100)
                end = min(len(text), match.end() + 100)
                snippet = text[start:end].replace('\n', ' ').strip()
                print(f"  Snippet: ... {snippet} ...")
            print("-" * 60)

def search_docx(filepath):
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            if "word/document.xml" in z.namelist():
                xml_content = z.read("word/document.xml").decode('utf-8', errors='ignore')
                # Strip XML tags
                text = re.sub(r'<[^>]+>', ' ', xml_content)
                search_text(text, os.path.basename(filepath))
    except Exception as e:
        print(f"Error reading docx {filepath}: {e}")

def search_xlsx(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        all_text = []
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                row_str = " ".join([str(val) for val in row if val is not None])
                if row_str:
                    all_text.append(row_str)
        text = " \n ".join(all_text)
        search_text(text, os.path.basename(filepath))
    except Exception as e:
        # print(f"Error reading xlsx {filepath}: {e}")
        pass

dirs = [r"D:\Desktop", r"D:\Desktop\VITAL_Scan", r"D:\Desktop\PORTFOLIO\PORTFOLIO - VITAL\data"]
for directory in dirs:
    if not os.path.exists(directory):
        continue
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        if not os.path.isfile(filepath):
            continue
        ext = os.path.splitext(filename)[1].lower()
        if ext == '.docx':
            search_docx(filepath)
        elif ext == '.xlsx':
            search_xlsx(filepath)
        elif ext == '.json':
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    search_text(content, filename)
            except Exception:
                pass
        elif ext == '.txt':
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    search_text(content, filename)
            except Exception:
                pass
