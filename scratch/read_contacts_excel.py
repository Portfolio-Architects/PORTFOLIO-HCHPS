import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\사업운영 관련 연락처.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Workbook loaded: {file_path}")
    print(f"Sheets: {wb.sheetnames}")
    print("=" * 60)
    
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"Sheet: {name}")
        rows = list(sheet.iter_rows(values_only=True))
        for r_idx, row in enumerate(rows[:30]):
            if any(cell is not None for cell in row):
                cleaned_row = [str(cell) if cell is not None else '' for cell in row[:10]]
                print(f"Row {r_idx+1:02d}: {cleaned_row}")
        print("=" * 60)
except Exception as e:
    print(f"Error: {e}")
