import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['사전사후측정결과(출석부 포함)']

print("Excel Columns and Row 13 Values:")
for col in range(1, sheet.max_column + 1):
    header_val1 = sheet.cell(row=11, column=col).value
    header_val2 = sheet.cell(row=12, column=col).value
    header = f"{header_val1 or ''} / {header_val2 or ''}".strip(" /")
    val = sheet.cell(row=13, column=col).value
    if val is not None or header:
        print(f"Col {col:03d} | Header: {header:<40} | Row 13 Val: {val}")
