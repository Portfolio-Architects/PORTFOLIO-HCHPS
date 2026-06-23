import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return None

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # ----------------------------------------------------
    # 1. Parse Metadata from 사전사후측정결과 sheet
    # ----------------------------------------------------
    sheet_data = wb['사전사후측정결과(출석부 포함)']
    
    district = sheet_data['B2'].value
    location = sheet_data['B3'].value
    period = sheet_data['B4'].value
    day_time = sheet_data['B5'].value
    registered_count = sheet_data['B6'].value
    instructor = sheet_data['B7'].value
    
    print("============================================================")
    print("■ 프로그램 개요 (Metadata)")
    print("============================================================")
    print(f"- 자치구: {district}")
    print(f"- 운영장소: {location}")
    print(f"- 운영기간: {period}")
    print(f"- 운영요일/시간: {day_time}")
    print(f"- 등록인원: {registered_count}")
    print(f"- 강사명/소속: {instructor}")
    print("============================================================")
    
    # ----------------------------------------------------
    # 2. Parse Participant Details
    # ----------------------------------------------------
    participants = []
    for row_idx in range(13, 212):
        row = [sheet_data.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3] # Column D
        if not name:
            continue
            
        serial_no = row[1] # Column B
        age = row[5] # Column F
        
        # Pre-measurements
        pre_height = safe_float(row[6])
        pre_weight = safe_float(row[7])
        pre_muscle = safe_float(row[8])
        pre_fat = safe_float(row[9])
        pre_bmi = safe_float(row[10])
        pre_waist = safe_float(row[11])
        
        # Post-measurements
        post_height = safe_float(row[12])
        post_weight = safe_float(row[13])
        post_muscle = safe_float(row[14])
        post_fat = safe_float(row[15])
        post_bmi = safe_float(row[16])
        post_waist = safe_float(row[17])
        
        # Attendance
        total_days = row[70] # BS (Col 71)
        attended_days = row[71] # BT (Col 72)
        attendance_rate = safe_float(row[72]) # BU (Col 73)
        
        participants.append({
            'no': serial_no,
            'name': name,
            'age': age,
            'pre': {'height': pre_height, 'weight': pre_weight, 'muscle': pre_muscle, 'fat': pre_fat, 'bmi': pre_bmi, 'waist': pre_waist},
            'post': {'height': post_height, 'weight': post_weight, 'muscle': post_muscle, 'fat': post_fat, 'bmi': post_bmi, 'waist': post_waist},
            'attendance': {'total': total_days, 'attended': attended_days, 'rate': attendance_rate}
        })
        
    print(f"\n■ 등록 회원 측정 및 출석 결과 (총 {len(participants)}명)")
    print("=" * 110)
    print(f"{'연번':<4}{'이름':<6}{'나이':<4}{'출석(율)':<12}{'사전 체중/근육/지방':<22}{'사후 체중/근육/지방':<22}{'체중 변화':<10}{'골격근 변화':<10}{'지방률 변화':<10}")
    print("-" * 110)
    for p in participants:
        rate = p['attendance']['rate']
        if rate is not None:
            # If the rate was stored as float (e.g. 0.85 means 85%), scale it if it is <= 1.0 and total_days > 1
            if rate <= 1.0 and (p['attendance']['total'] is not None and p['attendance']['total'] > 1):
                rate_pct = rate * 100.0
            else:
                rate_pct = rate
            att_str = f"{p['attendance']['attended']}/{p['attendance']['total']} ({rate_pct:.1f}%)"
        else:
            att_str = "N/A"
        
        pre_str = f"{p['pre']['weight'] or 0:.1f} / {p['pre']['muscle'] or 0:.1f} / {p['pre']['fat'] or 0:.1f}"
        post_str = f"{p['post']['weight'] or 0:.1f} / {p['post']['muscle'] or 0:.1f} / {p['post']['fat'] or 0:.1f}" if p['post']['weight'] else "측정없음"
        
        # Changes
        if p['post']['weight'] and p['pre']['weight']:
            w_diff = p['post']['weight'] - p['pre']['weight']
            w_diff_str = f"{w_diff:+.1f} kg"
        else:
            w_diff_str = "-"
            
        if p['post']['muscle'] and p['pre']['muscle']:
            m_diff = p['post']['muscle'] - p['pre']['muscle']
            m_diff_str = f"{m_diff:+.1f} kg"
        else:
            m_diff_str = "-"
            
        if p['post']['fat'] and p['pre']['fat']:
            f_diff = p['post']['fat'] - p['pre']['fat']
            f_diff_str = f"{f_diff:+.1f}%"
        else:
            f_diff_str = "-"
            
        print(f"{p['no']:<4}{p['name']:<6}{p['age']:<4}{att_str:<15}{pre_str:<25}{post_str:<25}{w_diff_str:<12}{m_diff_str:<12}{f_diff_str:<12}")
    print("=" * 110)
    
    # ----------------------------------------------------
    # 3. Parse Satisfaction Survey
    # ----------------------------------------------------
    sheet_survey = wb['만족도조사결과']
    survey_rows = list(sheet_survey.iter_rows(values_only=True))
    
    questions = [cell for cell in survey_rows[1] if cell is not None]
    
    respondents = []
    q_sums = [0.0] * 11
    q_counts = [0] * 11
    comments = []
    
    for row in survey_rows[2:]:
        if not row or row[0] is None:
            continue
        serial = row[0]
        scores = []
        for i in range(1, 12): # Q1 to Q11
            val = row[i]
            if val is not None:
                score = safe_float(val)
                if score is not None:
                    q_sums[i-1] += score
                    q_counts[i-1] += 1
                    scores.append(score)
                else:
                    scores.append(None)
            else:
                scores.append(None)
        
        comment = row[12] if len(row) > 12 else None
        respondents.append({
            'no': serial,
            'scores': scores,
            'comment': comment
        })
        if comment:
            comments.append(f"[{serial}번 회원] {comment.strip()}")
            
    print(f"\n■ 사후 만족도 조사 결과 (총 {len(respondents)}명 응답)")
    print("=" * 80)
    for idx in range(11):
        q_label = questions[idx+1]
        avg_score = q_sums[idx] / q_counts[idx] if q_counts[idx] > 0 else 0.0
        print(f"Q{idx+1:02d}. {q_label[:55]}... => 평균: {avg_score:.2f}점 / 5.00점 (응답 {q_counts[idx]}명)")
    
    print("\n■ 회원 기타 의견 및 건의 사항 (Q12)")
    print("-" * 80)
    for c in comments:
        print(c)
    print("=" * 80)

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
