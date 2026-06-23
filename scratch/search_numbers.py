import os
import sys
import openpyxl
import re

sys.stdout.reconfigure(encoding='utf-8')

print("Searching Desktop and Workspace for values 24.9 and 127...")

keywords = ['24.9', '127']

def search_text(text, filename):
    for kw in keywords:
        if kw in text:
            print(f"Match found for '{kw}' in file: {filename}")
            # print snippet
            idx = text.find(kw)
            print("  Snippet:", text[max(0, idx-100):min(len(text), idx+200)].replace('\n', ' ').strip())

for root, dirs, files in os.walk(r"d:\Desktop"):
    if ".git" in root or ".next" in root or "node_modules" in root:
        continue
    for file in files:
        filepath = os.path.join(root, file)
        ext = os.path.splitext(file)[1].lower()
        
        if ext in ['.txt', '.json', '.csv', '.md']:
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    search_text(content, filepath)
            except Exception:
                pass
        elif ext == '.xlsx':
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                        row_str = " ".join([str(val) for val in row if val is not None])
                        for kw in keywords:
                            if kw in row_str:
                                print(f"Match found for '{kw}' in XLSX: {filepath} (Sheet: {sheetname}, Row: {r_idx+1})")
                                print("  Row:", row_str[:200])
            except Exception:
                pass
