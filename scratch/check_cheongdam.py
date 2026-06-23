import sys
import openpyxl
import os

sys.stdout.reconfigure(encoding='utf-8')

desktop_dir = r"D:\Desktop"
cheongdam_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_청담평생학습관.xlsx")

try:
    wb = openpyxl.load_workbook(cheongdam_file, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    print("Scanning Cheongdam Lifelong Learning Center sheet...")
    completed_count = 0
    registered_count = 0
    for r in range(13, 212):
        name = sheet.cell(row=r, column=4).value
        if not name:
            continue
        registered_count += 1
        post_w = sheet.cell(row=r, column=14).value # Column N (post weight)
        if post_w is not None:
            completed_count += 1
            print(f"  Row {r:02d}: {name} completed post-test. (Post weight: {post_w})")
            
    print(f"Total registered: {registered_count}")
    print(f"Total completed: {completed_count}")
    
except Exception as e:
    print(f"Error: {e}")
