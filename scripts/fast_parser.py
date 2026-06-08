import sys
import os
import fitz  # PyMuPDF
import json
import itertools

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def parse_pdf(file_path):
    doc = fitz.open(file_path)
    text = []
    for page in doc:
        words = page.get_text('words')
        # sort by vertical line (rounded by 5) then horizontal position
        words.sort(key=lambda w: (round(w[3]/5), w[0]))
        for k, g in itertools.groupby(words, key=lambda w: round(w[3]/5)):
            line = ' '.join(w[4] for w in sorted(g, key=lambda w: w[0]))
            text.append(line)
    return '\n'.join(text)

def parse_xlsx(file_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        out = []
        for sheet_name in wb.sheetnames:
            out.append(f"=== Sheet: {sheet_name} ===")
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    line = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    out.append(line)
        return '\n'.join(out)
    except Exception as e:
        return f"Excel 파싱 오류: {str(e)}"

def parse_txt(file_path):
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()

def parse_hwpx(file_path):
    try:
        import zipfile
        import re
        import xml.etree.ElementTree as ET
        
        text_content = []
        with zipfile.ZipFile(file_path, 'r') as zf:
            sections = [f for f in zf.namelist() if f.startswith('Contents/section') and f.endswith('.xml')]
            sections.sort()
            
            for section in sections:
                xml_data = zf.read(section)
                try:
                    root = ET.fromstring(xml_data)
                    text_parts = []
                    for elem in root.iter():
                        # Extract text from tag endings with 't' (e.g. <hp:t>)
                        if elem.text and elem.tag.endswith('t'):
                            text_parts.append(elem.text)
                    if text_parts:
                        text_content.append(' '.join(text_parts))
                except Exception:
                    decoded = xml_data.decode('utf-8', errors='ignore')
                    cleaned = re.sub(r'<[^>]+>', ' ', decoded)
                    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
                    text_content.append(cleaned)
                    
        return '\n'.join(text_content)
    except Exception as e:
        return f"HWPX 파싱 오류: {str(e)}"

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "Missing file path parameter"}))
        sys.exit(1)
        
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(json.dumps({"success": False, "error": f"File not found: {file_path}"}))
        sys.exit(1)
        
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    try:
        if ext == '.pdf':
            content = parse_pdf(file_path)
        elif ext == '.hwpx':
            content = parse_hwpx(file_path)
        elif ext in ['.xlsx', '.xls']:
            content = parse_xlsx(file_path)
        elif ext in ['.txt', '.md', '.csv', '.json']:
            content = parse_txt(file_path)
        else:
            content = f"Unsupported file type: {ext}"
            
        print(json.dumps({"success": True, "content": content}, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}, ensure_ascii=False))
