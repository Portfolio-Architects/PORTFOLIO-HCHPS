import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['Sheet1']
    
    print("====================================================================")
    print("■ 7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx 기반 실제 출석 집계")
    print("====================================================================")
    print(f"{'연번':<4}{'이름':<6}{'1':<2}{'2':<2}{'3':<2}{'4':<2}{'5':<2}{'6':<2}{'7':<2}{'8':<2}{'9':<2}{'10':<3}{'11':<3}{'12':<3}{'13':<3}{'14':<3}{'15':<3}{'16':<3}{'출석일':<6}{'출석률':<8}")
    print("-" * 85)
    
    # Rows 4 to 15 have the 12 participants
    for r in range(4, 16):
        serial = sheet.cell(row=r, column=2).value
        name = sheet.cell(row=r, column=3).value
        if not name:
            continue
            
        sessions = []
        attended_count = 0
        for c in range(4, 20):  # columns D to S (1 to 16 sessions)
            val = sheet.cell(row=r, column=c).value
            # Count 1 as attended, others as not attended
            if val == 1 or val == '1':
                attended_count += 1
                sessions.append('O')
            elif val == 0 or val == '0':
                sessions.append('X')
            else:
                sessions.append('-') # No data recorded or absent
                
        rate = (attended_count / 16.0) * 100.0
        
        # Format session markers
        session_str = "".join(f"{s:<2}" for s in sessions)
        print(f"{serial:<4}{name:<6}{session_str}{attended_count:>3} / 16  {rate:>6.1f}%")
        
    print("====================================================================")
    
except Exception as e:
    print(f"Error: {e}")
