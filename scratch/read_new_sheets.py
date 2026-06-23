import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\사업운영 관련 연락처.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for name in ['사업운영', '기부관련', '양재천 행사관련']:
    if name in wb.sheetnames:
        print(f"=== Sheet: {name} ===")
        sheet = wb[name]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if any(cell is not None for cell in row):
                cleaned_row = [str(cell) if cell is not None else '' for cell in row[:10]]
                print(f"Row {r_idx+1:02d}: {cleaned_row}")
        print("=" * 60)
