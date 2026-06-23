import sys
import openpyxl
import os
import math

sys.stdout.reconfigure(encoding='utf-8')

desktop_dir = r"D:\Desktop"
chulbal_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx")
haemeoji_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx")

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return None

def calculate_paired_t_test(pre_list, post_list):
    n = len(pre_list)
    if n < 2:
        return None, None, None, None, None
        
    diffs = [post - pre for pre, post in zip(pre_list, post_list)]
    mean_diff = sum(diffs) / n
    
    # Standard deviation of differences
    sq_diff_sum = sum((d - mean_diff) ** 2 for d in diffs)
    std_dev = math.sqrt(sq_diff_sum / (n - 1))
    
    # Standard error
    std_err = std_dev / math.sqrt(n)
    
    # T-statistic
    if std_err == 0:
        t_stat = 0.0
    else:
        t_stat = mean_diff / std_err
        
    df = n - 1
    t_abs = abs(t_stat)
    
    try:
        theta = math.atan(t_abs / math.sqrt(df))
        if df == 1:
            p_val = 1.0 - (2.0 / math.pi) * theta
        elif df == 2:
            p_val = 1.0 - t_abs / math.sqrt(t_abs**2 + 2)
        elif df == 3:
            p_val = 1.0 - (2.0 / math.pi) * (theta + t_abs * math.sqrt(df) / (t_abs**2 + df))
        elif df == 4:
            x = t_abs / math.sqrt(t_abs**2 + 4)
            p_val = 2.0 * (1.0 - (0.5 + 0.5 * (x + 0.5 * x * (1.0 - x**2))))
        else:
            # Normal approximation for general df (very accurate for df >= 10)
            z = t_abs * (1.0 - 1.0 / (4.0 * df)) / math.sqrt(1.0 + t_abs**2 / (2.0 * df))
            p_val = math.erfc(z / math.sqrt(2.0))
    except Exception:
        p_val = None
        
    return mean_diff, std_dev, std_err, t_stat, p_val

try:
    combined_data = {
        'weight': ([], []),
        'muscle': ([], []),
        'fat': ([], []),
        'waist': ([], []),
        'gpaq': ([], [])
    }
    
    # List to hold details of all completers
    completers = []
    
    # 1. Load Chulbal Madang Completers (N=5)
    wb_chul = openpyxl.load_workbook(chulbal_file, data_only=True)
    sheet_chul = wb_chul['사전사후측정결과(출석부 포함)']
    for r in range(13, 212):
        row = [sheet_chul.cell(row=r, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3]
        if not name:
            continue
        post_w = safe_float(row[13])
        if post_w is not None:
            pre_w = safe_float(row[7])
            pre_m = safe_float(row[8])
            post_m = safe_float(row[14])
            pre_f = safe_float(row[9])
            post_f = safe_float(row[15])
            pre_waist = safe_float(row[11])
            post_waist = safe_float(row[17])
            gpaq_pre = safe_float(row[40]) if row[40] is not None else 0.0
            gpaq_post = safe_float(row[66]) if row[66] is not None else 0.0
            
            completers.append({
                'loc': '출발마당', 'name': name,
                'weight': (pre_w, post_w), 'muscle': (pre_m, post_m),
                'fat': (pre_f, post_f), 'waist': (pre_waist, post_waist),
                'gpaq': (gpaq_pre, gpaq_post)
            })
            
    # 2. Load Haemeoji Completers (N=9)
    wb_hae = openpyxl.load_workbook(haemeoji_file, data_only=True)
    sheet_hae = wb_hae['사전사후측정결과(출석부 포함)']
    for r in range(13, 212):
        row = [sheet_hae.cell(row=r, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3]
        if not name:
            continue
        post_w = safe_float(row[13])
        if post_w is not None:
            pre_w = safe_float(row[7])
            pre_m = safe_float(row[8])
            post_m = safe_float(row[14])
            pre_f = safe_float(row[9])
            post_f = safe_float(row[15])
            pre_waist = safe_float(row[11])
            post_waist = safe_float(row[17])
            gpaq_pre = safe_float(row[40]) if row[40] is not None else 0.0
            gpaq_post = safe_float(row[66]) if row[66] is not None else 0.0
            
            completers.append({
                'loc': '해맞이공원', 'name': name,
                'weight': (pre_w, post_w), 'muscle': (pre_m, post_m),
                'fat': (pre_f, post_f), 'waist': (pre_waist, post_waist),
                'gpaq': (gpaq_pre, gpaq_post)
            })
            
    # Populate combined_data arrays
    for c in completers:
        if c['weight'][0] is not None and c['weight'][1] is not None:
            combined_data['weight'][0].append(c['weight'][0])
            combined_data['weight'][1].append(c['weight'][1])
        if c['muscle'][0] is not None and c['muscle'][1] is not None:
            combined_data['muscle'][0].append(c['muscle'][0])
            combined_data['muscle'][1].append(c['muscle'][1])
        if c['fat'][0] is not None and c['fat'][1] is not None:
            combined_data['fat'][0].append(c['fat'][0])
            combined_data['fat'][1].append(c['fat'][1])
        if c['waist'][0] is not None and c['waist'][1] is not None:
            combined_data['waist'][0].append(c['waist'][0])
            combined_data['waist'][1].append(c['waist'][1])
        combined_data['gpaq'][0].append(c['gpaq'][0])
        combined_data['gpaq'][1].append(c['gpaq'][1])

    # 3. Print combined table
    print("==========================================================================================")
    print(f"■ 통합 대응표본 t-검정 (Combined Paired t-test) 결과 (N = {len(completers)}, df = {len(completers)-1})")
    print("==========================================================================================")
    print(f"{'평가 지표':<15}{'사전 평균':<12}{'사후 평균':<12}{'평균 변화량':<12}{'표준편차(diff)':<15}{'t-통계량':<12}{'p-value':<12}{'유의성 (p < 0.05)':<15}")
    print("-" * 105)
    
    for metric, (pre_vals, post_vals) in combined_data.items():
        avg_pre = sum(pre_vals) / len(pre_vals)
        avg_post = sum(post_vals) / len(post_vals)
        mean_diff, std_dev, std_err, t_stat, p_val = calculate_paired_t_test(pre_vals, post_vals)
        
        sig_str = "유의함 (Significant)" if p_val is not None and p_val < 0.05 else "유의하지 않음"
        p_val_str = f"{p_val:.4f}" if p_val is not None else "N/A"
        t_stat_str = f"{t_stat:+.3f}" if t_stat is not None else "N/A"
        mean_diff_str = f"{mean_diff:+.2f}"
        
        print(f"{metric.capitalize():<15}{avg_pre:<15.2f}{avg_post:<15.2f}{mean_diff_str:<15}{std_dev:<15.4f}{t_stat_str:<12}{p_val_str:<12}{sig_str:<15}")
    print("==========================================================================================")
    
    # 4. Check if there are correlations between attendance rate and weight / muscle changes
    # Let's read attendance rate for all 14 completers
    print("\n■ 출석일수와 신체 지표 변화량 간의 상관 관계 확인")
    print("-" * 80)
    print(f"{'위치':<10}{'이름':<6}{'출석일수':<8}{'체중 변화':<10}{'골격근 변화':<12}{'체지방률 변화':<12}")
    
    # We need attendance count for Chulbal
    chul_att = {}
    sheet_chul_att = wb_chul['사전사후측정결과(출석부 포함)']
    # Actually we can get it from 7. 체력향상... 출발마당
    wb_chul_att = openpyxl.load_workbook(os.path.join(desktop_dir, "7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx"), data_only=True)
    sheet_chul_att = wb_chul_att['Sheet1']
    for r in range(4, 16):
        name = sheet_chul_att.cell(row=r, column=3).value
        # map name
        if name == '전희정': name = '정희정'
        attended = 0
        for c in range(4, 20):
            val = sheet_chul_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended += 1
        chul_att[name] = attended
        
    # Attendance for Haemeoji
    hae_att = {}
    wb_hae_att = openpyxl.load_workbook(os.path.join(desktop_dir, "7. 체력향상 프로그램 출석부_2026_강남구(해맞이).xlsx"), data_only=True)
    sheet_hae_att = wb_hae_att[wb_hae_att.sheetnames[0]]
    for r in range(4, 40):
        name = sheet_hae_att.cell(row=r, column=3).value
        if not name:
            continue
        attended = 0
        for c in range(4, 20):
            val = sheet_hae_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended += 1
        hae_att[name] = attended
        
    for c in completers:
        att = chul_att.get(c['name'], 0) if c['loc'] == '출발마당' else hae_att.get(c['name'], 0)
        w_diff = c['weight'][1] - c['weight'][0]
        m_diff = c['muscle'][1] - c['muscle'][0]
        f_diff = c['fat'][1] - c['fat'][0]
        print(f"{c['loc']:<10}{c['name']:<6}{att:>2} / 16회   {w_diff:+.1f} kg    {m_diff:+.1f} kg      {f_diff:+.1f}%")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
