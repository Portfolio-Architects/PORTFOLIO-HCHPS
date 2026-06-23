import sys
import openpyxl
import os
import math

sys.stdout.reconfigure(encoding='utf-8')

desktop_dir = r"D:\Desktop"
chulbal_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx")
haemeoji_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx")

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return None

def pearson_correlation(x, y):
    n = len(x)
    if n < 2:
        return 0.0, 1.0
        
    mean_x = sum(x) / n
    mean_y = sum(y) / n
    
    num = sum((x_val - mean_x) * (y_val - mean_y) for x_val, y_val in zip(x, y))
    den_x = sum((x_val - mean_x) ** 2 for x_val in x)
    den_y = sum((y_val - mean_y) ** 2 for y_val in y)
    
    if den_x == 0 or den_y == 0:
        return 0.0, 1.0
        
    r = num / math.sqrt(den_x * den_y)
    
    # Calculate t-statistic for correlation
    # t = r * sqrt(n-2) / sqrt(1-r^2)
    if abs(r) == 1.0:
        t_stat = 9999.0
    else:
        t_stat = r * math.sqrt(n - 2) / math.sqrt(1 - r**2)
        
    df = n - 2
    t_abs = abs(t_stat)
    
    # p-value approximation for t-dist with df
    try:
        theta = math.atan(t_abs / math.sqrt(df))
        if df == 1:
            p_val = 1.0 - (2.0 / math.pi) * theta
        elif df == 2:
            p_val = 1.0 - t_abs / math.sqrt(t_abs**2 + 2)
        elif df == 3:
            p_val = 1.0 - (2.0 / math.pi) * (theta + t_abs * math.sqrt(df) / (t_abs**2 + df))
        elif df == 4:
            x_val = t_abs / math.sqrt(t_abs**2 + 4)
            p_val = 2.0 * (1.0 - (0.5 + 0.5 * (x_val + 0.5 * x_val * (1.0 - x_val**2))))
        else:
            z = t_abs * (1.0 - 1.0 / (4.0 * df)) / math.sqrt(1.0 + t_abs**2 / (2.0 * df))
            p_val = math.erfc(z / math.sqrt(2.0))
    except Exception:
        p_val = 1.0
        
    return r, p_val

try:
    completers = []
    
    # 1. Load Chulbal Madang Completers (N=5)
    wb_chul = openpyxl.load_workbook(chulbal_file, data_only=True)
    sheet_chul = wb_chul['사전사후측정결과(출석부 포함)']
    
    # Get attendance count
    chul_att = {}
    wb_chul_att = openpyxl.load_workbook(os.path.join(desktop_dir, "7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx"), data_only=True)
    sheet_chul_att = wb_chul_att['Sheet1']
    for r in range(4, 16):
        name = sheet_chul_att.cell(row=r, column=3).value
        if name == '전희정': name = '정희정'
        attended = 0
        for c in range(4, 20):
            val = sheet_chul_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended += 1
        chul_att[name] = attended
        
    for r in range(13, 212):
        row = [sheet_chul.cell(row=r, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3]
        if not name:
            continue
        post_w = safe_float(row[13])
        if post_w is not None:
            pre_w = safe_float(row[7])
            pre_m = safe_float(row[8])
            post_m = safe_float(row[14])
            pre_f = safe_float(row[9])
            post_f = safe_float(row[15])
            gpaq_pre = safe_float(row[40]) if row[40] is not None else 0.0
            gpaq_post = safe_float(row[66]) if row[66] is not None else 0.0
            
            completers.append({
                'name': name,
                'attendance': chul_att.get(name, 0),
                'w_diff': post_w - pre_w,
                'm_diff': post_m - pre_m,
                'f_diff': post_f - pre_f,
                'g_diff': gpaq_post - gpaq_pre
            })
            
    # 2. Load Haemeoji Completers (N=9)
    wb_hae = openpyxl.load_workbook(haemeoji_file, data_only=True)
    sheet_hae = wb_hae['사전사후측정결과(출석부 포함)']
    
    hae_att = {}
    wb_hae_att = openpyxl.load_workbook(os.path.join(desktop_dir, "7. 체력향상 프로그램 출석부_2026_강남구(해맞이).xlsx"), data_only=True)
    sheet_hae_att = wb_hae_att[wb_hae_att.sheetnames[0]]
    for r in range(4, 40):
        name = sheet_hae_att.cell(row=r, column=3).value
        if not name:
            continue
        attended = 0
        for c in range(4, 20):
            val = sheet_hae_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended += 1
        hae_att[name] = attended
        
    for r in range(13, 212):
        row = [sheet_hae.cell(row=r, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3]
        if not name:
            continue
        post_w = safe_float(row[13])
        if post_w is not None:
            pre_w = safe_float(row[7])
            pre_m = safe_float(row[8])
            post_m = safe_float(row[14])
            pre_f = safe_float(row[9])
            post_f = safe_float(row[15])
            gpaq_pre = safe_float(row[40]) if row[40] is not None else 0.0
            gpaq_post = safe_float(row[66]) if row[66] is not None else 0.0
            
            completers.append({
                'name': name,
                'attendance': hae_att.get(name, 0),
                'w_diff': post_w - pre_w,
                'm_diff': post_m - pre_m,
                'f_diff': post_f - pre_f,
                'g_diff': gpaq_post - gpaq_pre
            })
            
    # Calculate correlations
    att_list = [float(x['attendance']) for x in completers]
    w_diff_list = [x['w_diff'] for x in completers]
    m_diff_list = [x['m_diff'] for x in completers]
    f_diff_list = [x['f_diff'] for x in completers]
    g_diff_list = [x['g_diff'] for x in completers]
    
    print("====================================================================")
    print("■ 출석횟수(운동 횟수)와 신체변화 간의 피어슨 상관계수(r) 분석 (N=14)")
    print("====================================================================")
    
    r_w, p_w = pearson_correlation(att_list, w_diff_list)
    r_m, p_m = pearson_correlation(att_list, m_diff_list)
    r_f, p_f = pearson_correlation(att_list, f_diff_list)
    r_g, p_g = pearson_correlation(att_list, g_diff_list)
    
    print(f"1. 출석횟수 vs 체중 변화량:")
    print(f"   - 상관계수 r: {r_w:+.4f} (p-value: {p_w:.4f})")
    print(f"   - 해석: {'유의미함 (p < 0.05)' if p_w < 0.05 else '통계적 상관성 없음'}")
    
    print(f"\n2. 출석횟수 vs 골격근량 변화량:")
    print(f"   - 상관계수 r: {r_m:+.4f} (p-value: {p_m:.4f})")
    print(f"   - 해석: {'유의미함 (p < 0.05)' if p_m < 0.05 else '통계적 상관성 없음'}")
    
    print(f"\n3. 출석횟수 vs 체지방률 변화량:")
    print(f"   - 상관계수 r: {r_f:+.4f} (p-value: {p_f:.4f})")
    print(f"   - 해석: {'유의미함 (p < 0.05)' if p_f < 0.05 else '통계적 상관성 없음'}")
    
    print(f"\n4. 출석횟수 vs 신체활동량(GPAQ) 변화량:")
    print(f"   - 상관계수 r: {r_g:+.4f} (p-value: {p_g:.4f})")
    print(f"   - 해석: {'유의미함 (p < 0.05)' if p_g < 0.05 else '통계적 상관성 없음'}")
    print("====================================================================")
    
except Exception as e:
    print(f"Error: {e}")
