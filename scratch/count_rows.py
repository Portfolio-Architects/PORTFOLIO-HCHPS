import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\사업운영 관련 연락처.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for name in wb.sheetnames:
    sheet = wb[name]
    all_rows = list(sheet.iter_rows(values_only=True))
    non_empty_rows = [r for r in all_rows if any(cell is not None for cell in r)]
    print(f"Sheet: {name} | Total rows: {len(all_rows)} | Non-empty rows: {len(non_empty_rows)}")
