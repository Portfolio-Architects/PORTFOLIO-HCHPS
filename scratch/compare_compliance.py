import sys
import openpyxl
import os

sys.stdout.reconfigure(encoding='utf-8')

desktop_dir = r"D:\Desktop"
meas_file = os.path.join(desktop_dir, "4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx")
att_file = os.path.join(desktop_dir, "7. 체력향상 프로그램 출석부_2026_강남구(해맞이).xlsx")

try:
    wb_meas = openpyxl.load_workbook(meas_file, data_only=True)
    sheet_meas = wb_meas['사전사후측정결과(출석부 포함)']
    
    wb_att = openpyxl.load_workbook(att_file, data_only=True)
    sheet_att = wb_att[wb_att.sheetnames[0]]
    
    attendance = {}
    for r in range(4, 40):
        name = sheet_att.cell(row=r, column=3).value
        if not name:
            continue
        attended = 0
        for c in range(4, 20):
            val = sheet_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended += 1
        attendance[name] = attended
        
    high_compliance = []
    low_compliance = []
    
    for row_idx in range(13, 212):
        row = [sheet_meas.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 140)]
        name = row[3]
        if not name or name not in attendance:
            continue
        
        post_w = row[13]
        if post_w is None:
            continue # No post test
            
        pre_w = float(row[7])
        post_w = float(post_w)
        pre_m = float(row[8])
        post_m = float(row[14])
        pre_f = float(row[9])
        post_f = float(row[15])
        
        att_count = attendance[name]
        p_info = {
            'name': name,
            'attendance': att_count,
            'w_diff': post_w - pre_w,
            'm_diff': post_m - pre_m,
            'f_diff': post_f - pre_f
        }
        
        if att_count >= 10:
            high_compliance.append(p_info)
        else:
            low_compliance.append(p_info)
            
    print("====================================================================")
    print("■ 삼성 해맞이 공원: 출석 상태별(순응도) 신체 변화 비교")
    print("====================================================================")
    
    print(f"\n1. 고순응 그룹 (출석 10회 이상, N = {len(high_compliance)})")
    print("-" * 60)
    for p in high_compliance:
        print(f"  - {p['name']} (출석 {p['attendance']}회): 체중 {p['w_diff']:+.1f}kg | 골격근 {p['m_diff']:+.1f}kg | 체지방률 {p['f_diff']:+.1f}%")
    avg_w_high = sum(p['w_diff'] for p in high_compliance) / len(high_compliance) if high_compliance else 0
    avg_m_high = sum(p['m_diff'] for p in high_compliance) / len(high_compliance) if high_compliance else 0
    avg_f_high = sum(p['f_diff'] for p in high_compliance) / len(high_compliance) if high_compliance else 0
    print(f"  * 그룹 평균 변화: 체중 {avg_w_high:+.2f}kg | 골격근 {avg_m_high:+.2f}kg | 체지방률 {avg_f_high:+.2f}%")
    
    print(f"\n2. 저순응 그룹 (출석 10회 미만, N = {len(low_compliance)})")
    print("-" * 60)
    for p in low_compliance:
        print(f"  - {p['name']} (출석 {p['attendance']}회): 체중 {p['w_diff']:+.1f}kg | 골격근 {p['m_diff']:+.1f}kg | 체지방률 {p['f_diff']:+.1f}%")
    avg_w_low = sum(p['w_diff'] for p in low_compliance) / len(low_compliance) if low_compliance else 0
    avg_m_low = sum(p['m_diff'] for p in low_compliance) / len(low_compliance) if low_compliance else 0
    avg_f_low = sum(p['f_diff'] for p in low_compliance) / len(low_compliance) if low_compliance else 0
    print(f"  * 그룹 평균 변화: 체중 {avg_w_low:+.2f}kg | 골격근 {avg_m_low:+.2f}kg | 체지방률 {avg_f_low:+.2f}%")
    print("====================================================================")
    
except Exception as e:
    print(f"Error: {e}")
