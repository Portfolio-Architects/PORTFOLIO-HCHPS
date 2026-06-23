import openpyxl
import json
import os
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\사업운영 관련 연락처.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

contacts = []

def is_phone(val):
    if not val: return False
    cleaned = re.sub(r'[^0-9]', '', str(val))
    return len(cleaned) >= 7 and len(cleaned) <= 12

def clean_phone(val):
    if not val: return ""
    return str(val).strip().replace('\n', ', ')

# 1. 사업운영 시트 파싱
if '사업운영' in wb.sheetnames:
    sheet = wb['사업운영']
    last_org = ""
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        col3 = str(row[3]).strip() if row[3] is not None else ""
        col4 = str(row[4]).strip() if row[4] is not None else ""
        col5 = str(row[5]).strip() if row[5] is not None else ""
        col6 = str(row[6]).strip() if row[6] is not None else ""
        col8 = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
        col9 = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""
        
        if col0:
            last_org = col0
        org = last_org
        
        # 담당자1, 연락처1
        if col1 or col2:
            name = col1 if col1 else f"{org}(담당)"
            # 예외처리: 이름 란에 전화번호가 들어온 경우 스왑
            if is_phone(name) and not is_phone(col2):
                name, col2 = f"{org}(담당)", name
                
            email = col4 if '@' in col4 else ""
            notes_parts = []
            if org: notes_parts.append(f"소속: {org}")
            if col3: notes_parts.append(f"비고: {col3}")
            if col4 and not email: notes_parts.append(f"기타: {col4}")
            notes_parts.append("[출처: 사업운영]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(col2),
                "email": email,
                "notes": ", ".join(notes_parts)
            })
            
        # 담당자2, 연락처2
        if col5 or col6:
            # col6가 담당자2, col5가 연락처2 일 수 있음
            name = col6 if col6 else f"{org}(담당2)"
            phone = col5
            if is_phone(name) and not is_phone(phone):
                name, phone = f"{org}(담당2)", name
                
            notes_parts = []
            if org: notes_parts.append(f"소속: {org}")
            notes_parts.append("[출처: 사업운영]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(phone),
                "email": "",
                "notes": ", ".join(notes_parts)
            })
            
        # 특수열 (col 8, 9)
        if col8 or col9:
            # col8: 모두의 러너, col9: 전광하 대표
            name = col9 if col9 else f"{col8}(담당)"
            notes_parts = []
            if col8: notes_parts.append(f"소속: {col8}")
            notes_parts.append("[출처: 사업운영]")
            
            contacts.append({
                "name": name,
                "phone": "",
                "email": "",
                "notes": ", ".join(notes_parts)
            })
            
        # 고려대학교 최성규 (Row 3 특수케이스)
        # col4: 고려대학교 산학협력단 최성규, col5: 02-3407-4021
        if "최성규" in col4 and is_phone(col5):
            contacts.append({
                "name": "최성규",
                "phone": clean_phone(col5),
                "email": "",
                "notes": "소속: 고려대학교 산학협력단 [출처: 사업운영]"
            })

# 2. 기부관련 시트 파싱
if '기부관련' in wb.sheetnames:
    sheet = wb['기부관련']
    last_org = "서울사회복지공동모금회" # Default upper org
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        col3 = str(row[3]).strip() if row[3] is not None else ""
        col7 = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        col8 = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
        col9 = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""
        
        if col1:
            last_org = col1
        org = last_org
        
        # 성금/성품 담당 (col 8, col 9)
        if col8 or col9:
            name = col8 if col8 else f"{org}(기부담당)"
            notes_parts = []
            if org: notes_parts.append(f"소속: {org}")
            if col7: notes_parts.append(f"구분: {col7}")
            notes_parts.append("[출처: 기부관련]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(col9),
                "email": "",
                "notes": ", ".join(notes_parts)
            })
            
        # 단체 담당 (col 1, col 2, col 3)
        if col2 or col3:
            # col3이 전화번호인 경우에만
            if is_phone(col3):
                name = col2 if col2 else f"{org}(담당)"
                notes_parts = []
                if org: notes_parts.append(f"소속: {org}")
                notes_parts.append("[출처: 기부관련]")
                
                contacts.append({
                    "name": name,
                    "phone": clean_phone(col3),
                    "email": "",
                    "notes": ", ".join(notes_parts)
                })
            elif is_phone(col2) and not is_phone(col3):
                # 예: Row 12에서 col3은 없고 col2에 전화번호가 들어올 수 있음
                contacts.append({
                    "name": f"{org}(연락처)",
                    "phone": clean_phone(col2),
                    "email": "",
                    "notes": f"소속: {org} [출처: 기부관련]"
                })

# 3. 양재천 행사관련 시트 파싱
if '양재천 행사관련' in wb.sheetnames:
    sheet = wb['양재천 행사관련']
    last_org = ""
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        col3 = str(row[3]).strip() if row[3] is not None else ""
        col4 = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        col5 = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        col6 = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
        
        if col0:
            last_org = col0
        org = last_org
        
        # 담당자1, 연락처1
        if col1 or col2:
            name = col1 if col1 else f"{org}(담당)"
            # 이름 란에 전화번호가 들어온 경우 스왑
            if is_phone(name) and not is_phone(col2):
                name, col2 = f"{org}(담당)", name
                
            email = col3 if '@' in col3 else ""
            notes_parts = []
            if org: notes_parts.append(f"소속: {org}")
            if col3 and not email: notes_parts.append(f"비고: {col3}")
            notes_parts.append("[출처: 양재천 행사관련]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(col2),
                "email": email,
                "notes": ", ".join(notes_parts)
            })
            
        # 담당자2, 연락처2 (col 4, col 5, col 6)
        if col5 or col6:
            name = col5 if col5 else f"{col4}(담당)"
            notes_parts = []
            if col4: notes_parts.append(f"소속: {col4}")
            notes_parts.append("[출처: 양재천 행사관련]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(col6),
                "email": "",
                "notes": ", ".join(notes_parts)
            })

# 4. 감사패관련 시트 파싱
if '감사패관련' in wb.sheetnames:
    sheet = wb['감사패관련']
    last_org = ""
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        col3 = str(row[3]).strip() if row[3] is not None else ""
        col4 = str(row[4]).strip() if row[4] is not None else ""
        col5 = str(row[5]).strip() if row[5] is not None else ""
        
        if col0:
            last_org = col0
        org = last_org
        
        # 사람1
        if col2 or col1:
            name = col2 if col2 else org
            notes_parts = []
            if org: notes_parts.append(f"소속: {org}")
            notes_parts.append("[출처: 감사패관련]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(col1),
                "email": col3,
                "notes": ", ".join(notes_parts)
            })
            
        # 사람2
        if col5 or col4:
            name = col5 if col5 else org
            notes_parts = []
            if org: notes_parts.append(f"소속: {org}")
            notes_parts.append("[출처: 감사패관련]")
            
            contacts.append({
                "name": name,
                "phone": clean_phone(col4),
                "email": "",
                "notes": ", ".join(notes_parts)
            })

# 5. 건강증진지원실
if '건강증진지원실' in wb.sheetnames:
    contacts.append({
        "name": "홈페이지 관리자(업체)",
        "phone": "010-9431-0398",
        "email": "",
        "notes": "[출처: 건강증진지원실]"
    })

# 6. 어린이 성장발달 시스템
if '어린이 성장발달 시스템' in wb.sheetnames:
    sheet = wb['어린이 성장발달 시스템']
    
    contacts.append({
        "name": "신봉교(프로젝트 매니저)",
        "phone": "",
        "email": "devbeekei@kai-i.com",
        "notes": "소속: 카이아이컴퍼니 [출처: 어린이 성장발달 시스템]"
    })
    
    contacts.append({
        "name": "홍윤주",
        "phone": "",
        "email": "hyj9510@kai-i.com",
        "notes": "비고: 시스템 관련 문의는 메일로 소통 [출처: 어린이 성장발달 시스템]"
    })
    
    contacts.append({
        "name": "카이아이컴퍼니(대표)",
        "phone": "1670-2628",
        "email": "",
        "notes": "[출처: 어린이 성장발달 시스템]"
    })
    
    contacts.append({
        "name": "이주영",
        "phone": "010-4634-1821",
        "email": "",
        "notes": "소속: 카이아이컴퍼니, 역할: 시스템 문제 담당 [출처: 어린이 성장발달 시스템]"
    })
    
    contacts.append({
        "name": "오승준",
        "phone": "010-2462-7926",
        "email": "",
        "notes": "소속: 카이아이컴퍼니, 역할: 시스템 문제 담당 [출처: 어린이 성장발달 시스템]"
    })
    
    contacts.append({
        "name": "이은주",
        "phone": "010-2089-7757",
        "email": "",
        "notes": "소속: 카이아이컴퍼니, 역할: 시스템 문제 담당 [출처: 어린이 성장발달 시스템]"
    })

# 7. 지역사회건강조사 조사원
if '지역사회건강조사 조사원' in wb.sheetnames:
    sheet = wb['지역사회건강조사 조사원']
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx == 0:
            continue
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        col3 = str(row[3]).strip() if row[3] is not None else ""
        col4 = str(row[4]).strip() if row[4] is not None else ""
        col5 = str(row[5]).strip() if row[5] is not None else ""
        col6 = str(row[6]).strip() if row[6] is not None else ""
        
        if col0:
            notes_parts = []
            if col1: notes_parts.append(f"생년월일: {col1}")
            if col2: notes_parts.append(f"주소: {col2}")
            extra_notes = " / ".join(filter(None, [col5, col6]))
            if extra_notes: notes_parts.append(f"비고: {extra_notes}")
            notes_parts.append("[출처: 지역사회건강조사 조사원]")
            
            contacts.append({
                "name": col0,
                "phone": clean_phone(col3),
                "email": col4,
                "notes": ", ".join(notes_parts)
            })

# 8. 자문협력
if '자문협력' in wb.sheetnames:
    sheet = wb['자문협력']
    for row in sheet.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        col3 = str(row[3]).strip() if row[3] is not None else ""
        
        if col1:
            notes_parts = []
            if col0: notes_parts.append(f"소속: {col0}")
            if col3: notes_parts.append(f"비고: {col3}")
            notes_parts.append("[출처: 자문협력]")
            
            contacts.append({
                "name": col1,
                "phone": clean_phone(col2),
                "email": "",
                "notes": ", ".join(notes_parts)
            })

# 9. 면접위원
if '면접위원' in wb.sheetnames:
    sheet = wb['면접위원']
    for row in sheet.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if row[1] is not None else ""
        col2 = str(row[2]).strip() if row[2] is not None else ""
        
        if col1:
            notes_parts = []
            if col0: notes_parts.append(f"소속/직책: {col0}")
            notes_parts.append("[출처: 면접위원]")
            
            contacts.append({
                "name": col1,
                "phone": clean_phone(col2),
                "email": "",
                "notes": ", ".join(notes_parts)
            })

# output to scratch/extracted_contacts.json
output_file = r"D:\Desktop\PORTFOLIO\PORTFOLIO - VITAL\scratch\extracted_contacts.json"
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(contacts, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(contacts)} contacts to {output_file}")
