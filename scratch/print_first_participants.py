import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    for r in range(13, 16):
        name = sheet.cell(row=r, column=4).value # Col D
        serial = sheet.cell(row=r, column=2).value # Col B
        age = sheet.cell(row=r, column=6).value # Col F
        sex = sheet.cell(row=r, column=5).value # Col E
        pre_bmi = sheet.cell(row=r, column=11).value
        pre_waist = sheet.cell(row=r, column=12).value
        pre_bp = sheet.cell(row=r, column=108).value
        pre_gl = sheet.cell(row=r, column=110).value
        
        print(f"Row {r} | No.{serial} | Name: {name} | Age: {age} | Sex: {sex} | Pre-BMI: {pre_bmi} | Waist: {pre_waist} | BP: {pre_bp} | Glucose: {pre_gl}")
        # Print columns 100 to 120
        vals = {col: sheet.cell(row=r, column=col).value for col in range(100, 120) if sheet.cell(row=r, column=col).value is not None}
        print(f"  Cols 100-120: {vals}")
except Exception as e:
    print("Error:", e)
