import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Workbook successfully loaded: {file_path}")
    print(f"Sheet names: {wb.sheetnames}")
    print("=" * 60)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name} (Dimensions: {sheet.dimensions})")
        print("-" * 40)
        
        rows = list(sheet.iter_rows(values_only=True))
        # Print up to 40 rows to see everything
        for idx, row in enumerate(rows[:45]):
            if any(cell is not None for cell in row):
                cleaned_row = [str(cell)[:30] if cell is not None else '' for cell in row[:25]]
                print(f"Row {idx+1:02d}: {cleaned_row}")
        print("=" * 60)
except Exception as e:
    print(f"Error reading file: {e}")
