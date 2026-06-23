import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

files = [
    r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx",
    r"D:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_청담평생학습관.xlsx",
    r"D:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx"
]

print("Searching for participants with BP/Glucose data...")

for filepath in files:
    if not os.path.exists(filepath):
        continue
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb['사전사후측정결과(출석부 포함)']
        for r in range(13, 200):
            name = sheet.cell(row=r, column=4).value
            if not name:
                continue
            bp_sys = sheet.cell(row=r, column=108).value
            bp_dia = sheet.cell(row=r, column=109).value
            glucose = sheet.cell(row=r, column=110).value
            
            # Print if any of these are present
            if bp_sys is not None or bp_dia is not None or glucose is not None:
                serial = sheet.cell(row=r, column=2).value
                age = sheet.cell(row=r, column=6).value
                bmi = sheet.cell(row=r, column=11).value
                waist = sheet.cell(row=r, column=12).value
                bmi_str = f"{bmi:.2f}" if isinstance(bmi, (int, float)) else str(bmi)
                print(f"File: {os.path.basename(filepath)} | No.{serial} | Name: {name} | Age: {age} | BMI: {bmi_str} | Waist: {waist} | BP: {bp_sys}/{bp_dia} | Glucose: {glucose}")
    except Exception as e:
        print(f"Error reading {os.path.basename(filepath)}: {e}")
