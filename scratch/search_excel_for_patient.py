import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_files = [
    r"d:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_청담평생학습관.xlsx",
    r"d:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx",
    r"d:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx",
    r"d:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"
]

print("Searching Excel files for patient data...")

for filepath in excel_files:
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        continue
    print(f"Checking {filepath}...")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"  Sheet: {name} | Dimensions: {sheet.dimensions}")
            # Scan first 50 rows, 30 columns for patient name or any text
            for r in range(1, 100):
                row_vals = [sheet.cell(r, c).value for c in range(1, 50)]
                if any(v is not None for v in row_vals):
                    # Check if '김광숙' or other name keywords are in row
                    row_str = str(row_vals)
                    if '김광숙' in row_str or '광숙' in row_str or '광' in row_str:
                        print(f"    [FOUND IN ROW {r}]: {row_vals[:15]}")
                    elif r <= 20: # print first few rows to see headers
                        print(f"    Row {r}: {row_vals[:10]}")
    except Exception as e:
        print(f"  Error reading {filepath}: {e}")
