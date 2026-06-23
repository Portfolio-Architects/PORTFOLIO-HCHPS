import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

meas_file = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"
att_file = r"D:\Desktop\7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx"

# Mapping names from attendance sheet (source) to measurement sheet (target)
# Special mapping for '전희정' -> '정희정'
name_mapping = {
    '전희정': '정희정'
}

try:
    # 1. Load attendance data
    print("Loading attendance records...")
    wb_att = openpyxl.load_workbook(att_file, data_only=True)
    sheet_att = wb_att['Sheet1']
    
    attendance_data = {}
    for r in range(4, 16):
        name = sheet_att.cell(row=r, column=3).value
        if not name:
            continue
        
        # Count attendance
        attended_count = 0
        for c in range(4, 20):  # columns D to S (1 to 16 sessions)
            val = sheet_att.cell(row=r, column=c).value
            if val == 1 or val == '1':
                attended_count += 1
        
        # Apply name mapping if exists
        mapped_name = name_mapping.get(name, name)
        attendance_data[mapped_name] = attended_count
        print(f"  Attendance: {name} ({mapped_name}) -> {attended_count} / 16 sessions")

    # 2. Load measurement file (without data_only to preserve formulas)
    print("\nLoading measurement file...")
    wb_meas = openpyxl.load_workbook(meas_file, data_only=False)
    sheet_meas = wb_meas['사전사후측정결과(출석부 포함)']
    
    # 3. Update BS and BT columns
    updated_count = 0
    for r in range(13, 25):
        name = sheet_meas.cell(row=r, column=4).value
        if not name:
            continue
            
        if name in attendance_data:
            # Column BS (Col 71) = 전체교육일수
            sheet_meas.cell(row=r, column=71).value = 16
            # Column BT (Col 72) = 출석일수
            sheet_meas.cell(row=r, column=72).value = attendance_data[name]
            # Column BU (Col 73) = 출석률 formula
            sheet_meas.cell(row=r, column=73).value = f"=BT{r}/BS{r}*100"
            print(f"  Updating row {r} for {name}: BS=16, BT={attendance_data[name]}")
            updated_count += 1
        else:
            print(f"  Warning: {name} not found in attendance data!")
            
    # Save the updated measurement file
    wb_meas.save(meas_file)
    print(f"\nSuccessfully updated {updated_count} rows in {meas_file}")
    
except Exception as e:
    print(f"Error: {e}")
