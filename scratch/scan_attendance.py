import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    print("Scanning sheet for attendance-related cells or dates...")
    print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
    
    # Check row 8 to 12 for keywords
    for r in range(1, 15):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val and any(kwd in str(val) for kwd in ['출석', '결석', '일수', '회차', '4.', '5.', '6.']):
                print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {str(val)[:50]}")
                
    # Also check if there's any data in columns beyond BS (Col 71) for rows 13 to 24
    print("\nChecking participant rows for data in columns after BS (71):")
    for r in range(13, 25):
        name = sheet.cell(row=r, column=4).value
        if name:
            non_empty_cols = []
            for c in range(71, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    non_empty_cols.append(f"{openpyxl.utils.get_column_letter(c)}:{val}")
            print(f"Row {r:02d} ({name}): {', '.join(non_empty_cols[:10])} ...")
            
except Exception as e:
    print(f"Error: {e}")
