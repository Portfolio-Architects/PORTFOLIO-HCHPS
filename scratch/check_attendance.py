import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    print("Row | Name | Col 71 (BS) | Col 72 (BT) | Col 73 (BU)")
    print("-" * 50)
    for row_idx in range(13, 25):
        name = sheet.cell(row=row_idx, column=4).value
        if name:
            val_71 = sheet.cell(row=row_idx, column=71).value
            val_72 = sheet.cell(row=row_idx, column=72).value
            val_73 = sheet.cell(row=row_idx, column=73).value
            print(f"{row_idx:02d}  | {name:<6} | {val_71} | {val_72} | {val_73}")
            
except Exception as e:
    print(f"Error: {e}")
