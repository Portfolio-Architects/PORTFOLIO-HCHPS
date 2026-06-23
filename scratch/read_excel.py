import sys
import os

if sys.platform.startswith('win'):
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# First check if openpyxl is installed
try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed. Trying to install...")
    os.system("pip install openpyxl")
    import openpyxl

file_path = r"d:\Desktop\★ 20260619 위원요구및처리결과(복지문화보건)_최종.xlsx"

if not os.path.exists(file_path):
    print("File not found:", file_path)
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_str = " | ".join([str(val) for val in row if val is not None])
            if '강현섭' in row_str or '메디헬스' in row_str or '보건소' in row_str:
                print(f"Row {r_idx}: {row_str}")
except Exception as e:
    print("Error:", e)
