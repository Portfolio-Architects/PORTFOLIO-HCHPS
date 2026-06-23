import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

keywords = ['체격', '신체', '측정', '혈압', '공복혈당', 'CVA', '머리척추각']

def extract_excel_data(file_path):
    print(f"\n==========================================")
    print(f"File: {os.path.basename(file_path)}")
    print(f"==========================================")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Find sheet names
        sheet_names = wb.sheetnames
        print("Sheets:", sheet_names)
        
        sheet = wb['사전사후측정결과(출석부 포함)']
        # Print first few columns and rows
        print("Row 12 columns:")
        header = [sheet.cell(row=12, column=col).value for col in range(1, 40)]
        print(header[:20])
        
        print("\nParticipants list:")
        count = 0
        for r in range(13, 200):
            name = sheet.cell(row=r, column=4).value # Col D
            if not name:
                continue
            serial = sheet.cell(row=r, column=2).value # Col B
            age = sheet.cell(row=r, column=6).value # Col F
            sex = sheet.cell(row=r, column=5).value # Col E (assuming)
            
            # Let's inspect BMI, Waist, Blood Pressure, CVA
            # pre-measurements
            pre_w = sheet.cell(row=r, column=8).value
            pre_bmi = sheet.cell(row=r, column=11).value
            pre_waist = sheet.cell(row=r, column=12).value
            
            # Blood Pressure / Glucose (Cols 108, 109, 110, etc.)
            pre_bp = sheet.cell(row=r, column=108).value
            pre_gl = sheet.cell(row=r, column=110).value
            
            # Let's see if there is CVA (Craniovertebral angle)
            # We don't know the exact column, let's scan all cells in this row for numeric values and print them if name is matched
            row_vals = [sheet.cell(row=r, column=col).value for col in range(1, 140)]
            print(f"No.{serial} | Name: {name} | Age: {age} | Sex: {sex} | Pre-BMI: {pre_bmi} | Waist: {pre_waist} | BP: {pre_bp} | Glucose: {pre_gl}")
            
            # Let's print non-empty values after column 100
            non_empty_cols = {col: row_vals[col-1] for col in range(100, len(row_vals)+1) if col-1 < len(row_vals) and row_vals[col-1] is not None}
            if non_empty_cols:
                print(f"  Col 100+ data: {non_empty_cols}")
            
            count += 1
            if count > 20:
                print("  ... showing first 20 participants")
                break
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

files = [
    r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx",
    r"D:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_청담평생학습관.xlsx",
    r"D:\Desktop\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx"
]

for f in files:
    if os.path.exists(f):
        extract_excel_data(f)
