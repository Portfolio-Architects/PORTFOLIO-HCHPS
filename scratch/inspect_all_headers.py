import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    print("ALL Row 11 headers:")
    for col in range(1, 141):
        letter = openpyxl.utils.get_column_letter(col)
        h11 = sheet.cell(row=11, column=col).value
        h10 = sheet.cell(row=10, column=col).value
        h8 = sheet.cell(row=8, column=col).value
        
        # If there's any value in row 8, 10 or 11, let's print it
        if h11 or h10 or h8:
            print(f"Col {col:03d} ({letter}): Row8='{h8}', Row10='{h10}', Row11='{h11}'")
            
except Exception as e:
    print(f"Error: {e}")
