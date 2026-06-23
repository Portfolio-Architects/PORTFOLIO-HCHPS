import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"D:\Desktop\VITAL_Scan\4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx"

try:
    # Load workbook WITHOUT data_only=True to see formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['사전사후측정결과(출석부 포함)']
    
    print("--- Checking cell values/formulas for Rows 13-15 ---")
    for r in [13, 14, 15]:
        name = sheet.cell(row=r, column=4).value
        formula_bs = sheet.cell(row=r, column=71).value  # BS
        formula_bt = sheet.cell(row=r, column=72).value  # BT
        formula_bu = sheet.cell(row=r, column=73).value  # BU
        print(f"Row {r:02d} ({name}):")
        print(f"  BS (전체교육일수): {repr(formula_bs)}")
        print(f"  BT (출석일수):     {repr(formula_bt)}")
        print(f"  BU (출석률(%)):   {repr(formula_bu)}")
        
except Exception as e:
    print(f"Error: {e}")
