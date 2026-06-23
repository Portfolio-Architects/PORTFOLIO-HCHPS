import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

import os

desktop_dir = r"D:\Desktop"
locations = [
    {
        'name': '출발마당',
        'meas_file': '4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_출발마당.xlsx',
        'att_file': '7. 체력향상 프로그램 출석부_2026_강남구(출발마당).xlsx'
    },
    {
        'name': '청담평생',
        'meas_file': '4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_청담평생학습관.xlsx',
        'att_file': '7. 체력향상 프로그램 출석부_2026_강남구(창담평생학습관).xlsx'
    },
    {
        'name': '해맞이공원',
        'meas_file': '4. 측정 데이터 입력서식(측정결과, 출석부, 만족도 조사)_해맞이공원.xlsx',
        'att_file': '7. 체력향상 프로그램 출석부_2026_강남구(해맞이).xlsx'
    }
]

for loc in locations:
    print(f"\n============================================================")
    print(f"■ 위치: {loc['name']}")
    print(f"============================================================")
    
    # Check Measurement file
    meas_path = os.path.join(desktop_dir, loc['meas_file'])
    if os.path.exists(meas_path):
        try:
            wb = openpyxl.load_workbook(meas_path, data_only=True)
            sheet = wb['사전사후측정결과(출석부 포함)']
            print(f"Measurement File: {loc['meas_file']}")
            print(f"  District: {sheet['B2'].value}")
            print(f"  Location: {sheet['B3'].value}")
            print(f"  Period: {sheet['B4'].value}")
            print(f"  Days/Time: {sheet['B5'].value}")
            print(f"  Instructor: {sheet['B7'].value}")
            
            # Print BS, BT, BU values for the first 3 rows
            print("  First 3 participants BS, BT, BU values:")
            for r in [13, 14, 15]:
                name = sheet.cell(row=r, column=4).value
                if name:
                    bs = sheet.cell(row=r, column=71).value
                    bt = sheet.cell(row=r, column=72).value
                    bu = sheet.cell(row=r, column=73).value
                    print(f"    - {name}: BS={bs}, BT={bt}, BU={bu}")
        except Exception as e:
            print(f"  Error reading measurement file: {e}")
    else:
        print(f"  Measurement file NOT found: {meas_path}")
        
    # Check Attendance file
    att_path = os.path.join(desktop_dir, loc['att_file'])
    if os.path.exists(att_path):
        try:
            wb = openpyxl.load_workbook(att_path, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            print(f"Attendance File: {loc['att_file']}")
            # Count columns in row 2 (which has the sessions like 1회차, 2회차...)
            sessions = []
            for col in range(4, sheet.max_column + 1):
                val = sheet.cell(row=2, column=col).value
                if val:
                    sessions.append(val)
            print(f"  Sessions Count in Row 2: {len(sessions)} ({', '.join(sessions[:5])} ... {sessions[-1] if sessions else ''})")
            
            # Print first 3 rows of attendance data
            print("  First 3 participants attendance row:")
            for r in range(4, 7):
                name = sheet.cell(row=r, column=3).value
                if name:
                    att_vals = [sheet.cell(row=r, column=c).value for c in range(4, sheet.max_column + 1)]
                    print(f"    - {name}: {att_vals}")
        except Exception as e:
            print(f"  Error reading attendance file: {e}")
    else:
        print(f"  Attendance file NOT found: {att_path}")
